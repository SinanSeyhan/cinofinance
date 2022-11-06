import os
from unittest import result
import pandas as pd
from openpyxl import load_workbook
import yfinance as yf
from datetime import datetime


class Portfolio():
    def __init__(self) -> None:
        my_path = os.path.abspath(os.path.dirname(__file__))
        self.PORTFOLIO_PATH = os.path.join(my_path, '../raw_data/portfolio.xlsx')
        self.TICKER_PATH_TOTAL = os.path.join(my_path, '../raw_data/nasdaq_screener.csv')

    def get_portfolio(self):
        '''
        Function to read a excel file that contains Tickers and Lot amount per month.
        '''
        portfolio = []
        # Load the workbook, from the filename, setting read_only to False
        wb = load_workbook(filename=self.PORTFOLIO_PATH, read_only=False, keep_vba=False, data_only=True, keep_links=False)

        # Initialize the dictionary of tables
        tables_dict = {}

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:

            ws = wb[ws_name]

            # Get each table in the worksheet
            for tbl in ws.tables.values():
                # First, add some info about the table to the dictionary
                tables_dict[tbl.name] = {
                        'table_name': tbl.name,
                        'worksheet': ws_name,
                        'num_cols': len(tbl.tableColumns),
                        'table_range': tbl.ref}

                # Grab the 'data' from the table
                data = ws[tbl.ref]

                # Now convert the table 'data' to a Pandas DataFrame
                # First get a list of all rows, including the first header row
                rows_list = []
                for row in data:
                    # Get a list of all columns in each row
                    cols = []
                    for col in row:
                        cols.append(col.value)
                    rows_list.append(cols)

                # Create a pandas dataframe from the rows_list.
                # The first row is the column names
                df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

                # Add the dataframe to the dictionary of tables
                tables_dict[tbl.name]['dataframe'] = df
        # a dictionary of all tables in the Excel workbook
        portfolio = tables_dict['Table1']['dataframe']
        portfolio = portfolio.set_index('Tickers').dropna(axis='columns', how='all').T.reset_index()
        portfolio['index'] = pd.to_datetime(portfolio['index'], format='%Y-%m').dt.strftime('%Y-%m')
        portfolio = portfolio.set_index('index').T
        # Read the Stock information:
        ticker = pd.read_csv(self.TICKER_PATH_TOTAL).set_index('Symbol')
        ticker.drop(columns=['Last Sale', 'Net Change', '% Change', 'Market Cap',
                            'IPO Year', 'Volume'], axis=1, inplace=True)
        # Merge it with the Portfolio:
        portfolio = ticker.merge(portfolio, how='right', left_index=True, right_index=True)
        return portfolio


    def get_portfolio_data(self):
        portfolio = self.get_portfolio()
        tickers = list(portfolio.index)
        # Download the Portfolio data from yfinance:
        data = yf.download(tickers, period='1y', group_by='ticker', actions=True)
        return data

    def get_portfolio_dividend(self):
        portfolio = self.get_portfolio()
        data = self.get_portfolio_data()

        # Get the last 12 Month Dividends of the Portfolio:
        dividend = {}
        for tick in portfolio.index:
            dividend[tick] = data[tick]['Dividends']
        dividends = pd.DataFrame(dividend).reset_index()
        dividends['Date'] = dividends['Date'] + pd.DateOffset(days=20)
        dividends.Date = dividends.Date.dt.strftime('%Y-%m')
        dividends = dividends.groupby(by='Date').sum().T

        # Multiply dividends with portfolio
        result = portfolio.mul(dividends).dropna(axis=1)
        result['Name'] = portfolio['Name']
        return result


if __name__=='__main__':
    df = Portfolio().get_portfolio_dividend()
    print(df)
