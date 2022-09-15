from calendar import month
import pandas as pd
import datetime as dt
from datetime import datetime
import os
import numpy as np

from openpyxl import load_workbook

import yfinance as yf
from forex_python.converter import CurrencyRates


class Dividend():
    def __init__(self):

        my_path = os.path.abspath(os.path.dirname(__file__))
        global TICKER_PATH_TOTAL
        self.TICKER_PATH_TOTAL = os.path.join(my_path, '../raw_data/nasdaq_screener.csv')
        self.POOL_PATH = os.path.join(my_path, '../raw_data/pool.csv')
        self.POOL_FINAL = os.path.join(my_path, '../raw_data/pool_final.csv')
        self.PORTFOLIO_PATH = os.path.join(my_path, '../raw_data/portfolio.xlsx')

        # Pool for Stock picking
        tickers = pd.read_csv(self.POOL_PATH).to_dict('list')['Tickers']


        self.pool = dict(zip(tickers, [1] * len(tickers)))

    def usd2eur(self, amount):
        '''
        Function that converts US Dollars to EUR

        amount: float, US Dollars to convert
        '''
        return amount * CurrencyRates().get_rates('USD')['EUR']

    def get_portfolio(self):
        global portfolio
        portfolio = []
        # Load the workbook, from the filename, setting read_only to False
        wb = load_workbook(filename='../raw_data/portfolio.xlsx', read_only=False, keep_vba=False, data_only=True, keep_links=False)

        # Initialize the dictionary of tables
        tables_dict = {}

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:

            ws = wb[ws_name]

            # Get each table in the worksheet
            for tbl in ws.tables.values():
                # First, add some info about the table to the dictionary
                tables_dict[tbl.name] = {
                        'table_name': tbl.name,
                        'worksheet': ws_name,
                        'num_cols': len(tbl.tableColumns),
                        'table_range': tbl.ref}

                # Grab the 'data' from the table
                data = ws[tbl.ref]

                # Now convert the table 'data' to a Pandas DataFrame
                # First get a list of all rows, including the first header row
                rows_list = []
                for row in data:
                    # Get a list of all columns in each row
                    cols = []
                    for col in row:
                        cols.append(col.value)
                    rows_list.append(cols)

                # Create a pandas dataframe from the rows_list.
                # The first row is the column names
                df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

                # Add the dataframe to the dictionary of tables
                tables_dict[tbl.name]['dataframe'] = df
        # a dictionary of all tables in the Excel workbook
        portfolio = tables_dict['Table1']['dataframe']

        return portfolio


    def get_dividends(self, portfolio):
        '''
        Function that gives you a dataframe of the dividends, names and ROI of passed dictionary.

        portfolio: dict, with tickers as keys and lot size as values as user input.
        '''
        global month_columns
        total_columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Price']
        month_columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        brutto2netto = 0.855

        # Load the portfolio
        df = pd.DataFrame([portfolio]).T
        df.rename({0:'Lot'}, axis=1, inplace=True)

        # get the names of stocks
        df_tickers = pd.read_csv(self.TICKER_PATH_TOTAL)
        names = df_tickers[df_tickers['Symbol'].isin(list(portfolio.keys()))].iloc[:, 0:2].set_index('Symbol')
        df = df.join(names)

        # Get the Dividends
        divs = []
        for key, value in portfolio.items():
            temp = yf.Ticker(key)
            temp.history(period='1y')
            divs.append(temp.dividends) # * value)) --> For getting the lot size
        df2 = pd.DataFrame(divs, index=portfolio)
        df2.columns = df2.columns.month
        df2 = df2.groupby(df2.columns, axis=1).sum()
        df2.columns = [dt.date(1900,i,1).strftime('%b') for i in df2.columns]
        # Merge the dataframes
        result = df.merge(df2, left_index=True, right_index=True)

        #get price information
        price = [yf.download(i, period='1d')['Adj Close'].values[0] for i in portfolio.keys()]
        result['Price'] = price

        # Convert to EUR
        result[total_columns] = result[total_columns].apply(self.usd2eur)

        result[month_columns] = result[month_columns] * brutto2netto

        # Add annual Dividends and ROI
        result['Dividend, annual'] = result[month_columns].sum(axis=1)
        result['ROI, annual'] = result['Dividend, annual'] / result['Price']

        # Sort the dataframe
        result = result.sort_values('ROI, annual', ascending=False)

        return round(result, 5)

    def get_portfolio_dividends(self, portfolio):
        '''
        Gives a dataframe that is multiplicated with Lot size.

        portfolio: Dictionary
        '''
        df = self.get_dividends(portfolio)


        # Multiplying the Lot Size
        df[month_columns] = df[month_columns].multiply(df['Lot'], axis='index')
        df['Dividend, annual'] = df[month_columns].sum(axis=1)

        # Total Row
        df.loc["Total"] = df.sum(numeric_only=True)
        df.loc["Total", 'Price'] = 0
        df.loc["Total", 'Lot'] = 0
        df.loc["Total", 'ROI, annual'] = 0
        df.fillna(0, inplace=True)

        # Total income of the portfolio
        total = round(df.loc['Total', 'Dividend, annual'], 2)
        print(' ')
        print('#'*70)
        print(f'# The Total Annual Dividend Income of the portfolio is: {total} EUR     #')
        print('#'*70)
        print(' ')

        # Total Montly income
        current_month = datetime.now().strftime('%h')
        monthly = round(df.loc['Total', current_month], 2)
        print(' ')
        print('#'*70)
        print(f'# The Total Dividend Income for {current_month} of the portfolio is: {monthly} EUR      #')
        print('#'*70)
        print(' ')


        return df



    def get_dividend_pool(self, capital):

        '''
        Gives a dataframe from the pool
        '''
        df = self.get_dividends(self.pool)
        df['Lot'] = (capital / df['Price'])
        df['Lot'] = df['Lot'].apply(np.floor)

        month_columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        df[month_columns] = df[month_columns].multiply(df['Lot'], axis='index')
        df['Dividend, annual'] = df[month_columns].sum(axis=1)


        df.sort_values('Dividend, annual', ascending=False)
        df.to_csv(self.POOL_FINAL)
        return df

    def get_final_pool(self):
        df = pd.read_csv(self.POOL_FINAL)
        return df

if __name__=='__main__':
    # portfolio = {
    #             'NVDA': 2, # NVIDIA
    #             'STAG': 9, # STAG
    #             'MCO': 1, # MOODY'S
    #             'MSFT': 1, # MICROSOFT
    #             'AGNC': 20, # AGNC INVESTMENT
    #             'V': 1, # VISA
    #             'BNTX': 1, # BIONTECH
    #             }

    # df = Dividend().get_portfolio_dividends(portfolio)
    # print(df)
    df = Dividend().get_dividend_pool()
    print(df.columns)
